#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, json, math
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from universe_loader import build_universe

ROOT = Path(__file__).resolve().parent

def rsi(close, length=14):
    delta=close.diff(); gain=delta.clip(lower=0).ewm(alpha=1/length,adjust=False).mean(); loss=(-delta.clip(upper=0)).ewm(alpha=1/length,adjust=False).mean()
    return (100-100/(1+gain/loss.replace(0,np.nan))).fillna(50)
def atr(df, length=14):
    prev=df.Close.shift(); tr=pd.concat([df.High-df.Low,(df.High-prev).abs(),(df.Low-prev).abs()],axis=1).max(axis=1); return tr.ewm(alpha=1/length,adjust=False).mean()
def indicators(df):
    x=df.copy(); x['RSI']=rsi(x.Close); x['ATR']=atr(x); x['EMA9']=x.Close.ewm(span=9,adjust=False).mean(); x['EMA20']=x.Close.ewm(span=20,adjust=False).mean(); x['RVOL']=x.Volume/x.Volume.rolling(20,min_periods=5).mean().replace(0,np.nan); return x
def pivot_indices(series, side, left, right):
    v=series.to_numpy(float); out=[]
    for i in range(left,len(v)-right):
        window=v[i-left:i+right+1]
        if side=='low' and v[i]==np.nanmin(window):out.append(i)
        if side=='high' and v[i]==np.nanmax(window):out.append(i)
    return out
def score_setup(x,a,b,direction):
    av=float(x.ATR.iloc[b]); col='Low' if direction=='Bullish' else 'High'
    if not np.isfinite(av) or av<=0:return 0
    return round(15+min(25,abs(x.RSI.iloc[b]-x.RSI.iloc[a])*2)+min(20,abs(x[col].iloc[b]-x[col].iloc[a])/av*8)+min(15,5+(b-a)/4)+min(10,float(x.RVOL.iloc[b] if pd.notna(x.RVOL.iloc[b]) else 0)*5)+(15 if ((direction=='Bullish' and x.Close.iloc[b]>=x.EMA20.iloc[b]) or (direction=='Bearish' and x.Close.iloc[b]<=x.EMA20.iloc[b])) else 7.5),1)
def scan_one(df,cfg):
    x=indicators(df); c=cfg['signal']; setups=[]; raw=0
    for side,direction,col in [('low','Bullish','Low'),('high','Bearish','High')]:
        ps=pivot_indices(x[col],side,c['pivot_left'],c['pivot_right'])
        for j in range(1,len(ps)):
            a,b=ps[j-1],ps[j]; si=b+c['pivot_right']; gap=b-a
            if gap<c['min_pivot_gap'] or gap>c['max_pivot_gap'] or si>=len(x):continue
            p1,p2=float(x[col].iloc[a]),float(x[col].iloc[b]); r1,r2=float(x.RSI.iloc[a]),float(x.RSI.iloc[b])
            ok=(direction=='Bullish' and p2<p1 and r2>r1 and min(r1,r2)<=c['oversold']) or (direction=='Bearish' and p2>p1 and r2<r1 and max(r1,r2)>=c['overbought'])
            if not ok:continue
            raw+=1; setups.append({'direction':direction,'score':score_setup(x,a,b,direction),'signal_index':si,'pivot2_index':b,'price2':p2,'signal_date':x.index[si],'pivot2_date':x.index[b],'age':len(x)-1-si})
    score_passed=sum(s['score']>=c['minimum_score'] for s in setups)
    active=[s for s in setups if s['score']>=c['minimum_score'] and s['age']<=c['search_bars']]
    if not active:return None,raw,score_passed
    s=max(active,key=lambda z:(z['signal_index'],z['score'])); last=len(x)-1; ti=None
    if (s['direction']=='Bullish' and x.Close.iloc[-1]<s['price2']) or (s['direction']=='Bearish' and x.Close.iloc[-1]>s['price2']):status='INVALIDATED'
    elif s['direction']=='Bearish':
        ti=next((i for i in range(max(s['signal_index']+1,1),last+1) if x.Close.iloc[i]<x.EMA9.iloc[i] and x.Close.iloc[i-1]>=x.EMA9.iloc[i-1] and x.Close.iloc[i]<x.Close.iloc[i-1]),None)
        status='ENTRY_READY' if ti==last else ('TRIGGERED_PREVIOUSLY' if ti is not None else 'WAIT_EMA9')
    else:
        bi=next((i for i in range(s['signal_index']+1,last+1) if len(x.iloc[max(0,i-5):i])>=3 and x.Close.iloc[i]>x.iloc[max(0,i-5):i].High.max()),None)
        if bi is None:status='WAIT_STRUCTURE'
        else:
            for i in range(bi+1,last+1):
                low,high=sorted((float(x.EMA9.iloc[i]),float(x.EMA20.iloc[i]))); touched=x.Low.iloc[i]<=high and x.High.iloc[i]>=low
                if touched and x.Close.iloc[i]>x.Open.iloc[i] and x.Close.iloc[i]>x.Close.iloc[i-1] and x.Close.iloc[i]>x.EMA9.iloc[i]:ti=i;break
            status='ENTRY_READY' if ti==last else ('TRIGGERED_PREVIOUSLY' if ti is not None else 'WAIT_PULLBACK')
    delay=ti-s['signal_index'] if ti is not None else pd.NA
    s.update({'status':status,'trigger_delay':delay,'quality':'PRIME' if pd.notna(delay) and 6<=delay<=15 else 'STANDARD','latest_close':float(x.Close.iloc[-1]),'rsi':float(x.RSI.iloc[-1]),'ema9':float(x.EMA9.iloc[-1]),'ema20':float(x.EMA20.iloc[-1]),'atr':float(x.ATR.iloc[-1])})
    return s,raw,score_passed
def download(symbols,minimum,issues):
    import yfinance as yf
    out={}
    for start in range(0,len(symbols),50):
        batch=symbols[start:start+50]
        try:
            raw=yf.download(batch,period='10y',auto_adjust=True,group_by='ticker',threads=True,progress=False)
            for symbol in batch:
                try:
                    frame=(raw[symbol] if len(batch)>1 else raw)[['Open','High','Low','Close','Volume']].dropna(subset=['Open','High','Low','Close'])
                    if len(frame)>=minimum:out[symbol]=frame
                    else:issues.append(f'{symbol}: insufficient history {len(frame)}')
                except Exception as exc:issues.append(f'{symbol}: extraction failed: {exc}')
        except Exception as exc:issues.append(f'batch download failed: {exc}')
    return out
def trade_plan(signal,cfg):
    slip=cfg['risk']['slippage_bps']/10000; sign=1 if signal['direction']=='Bullish' else -1; entry=signal['latest_close']*(1+slip if sign==1 else 1-slip); risk=max(signal['atr'],entry*.02); stop=entry-sign*risk; amount=cfg['risk']['starting_capital']*cfg['risk']['risk_per_trade_pct']/100
    return {'Planned Entry':round(entry,4),'Stop':round(stop,4),'TP1':round(entry+sign*risk,4),'TP2':round(entry+sign*2*risk,4),'TP3':round(entry+sign*3*risk,4),'Risk Amount':round(amount,2),'Quantity':math.floor(amount/risk)}
def write_report(path,rows,no_setup,audit,universe,issues,funnel):
    expired=[r for r in rows if r['Status'] in {'INVALIDATED','EXPIRED','TRIGGERED_PREVIOUSLY'}]
    sheets={'Daily Signal Board':pd.DataFrame(rows),'Paper Decision Board':pd.DataFrame(rows),'Entry Ready':pd.DataFrame([r for r in rows if r['Status']=='ENTRY_READY']),'Waiting Setups':pd.DataFrame([r for r in rows if r['Status'].startswith('WAIT')]),'Expired Invalidated':pd.DataFrame(expired),'Event Funnel':pd.DataFrame(list(funnel.items()),columns=['Stage','Count']),'Universe Audit':audit,'Universe Health':universe,'No Setup Universe':pd.DataFrame(no_setup),'Data Issues':pd.DataFrame({'Issue':issues})}
    path.parent.mkdir(parents=True,exist_ok=True)
    with pd.ExcelWriter(path,engine='openpyxl') as writer:
        for name,frame in sheets.items():frame.to_excel(writer,sheet_name=name,index=False)
    wb=load_workbook(path)
    date_headers={'Signal Date'}
    for ws in wb.worksheets:
        ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
        headers={c.value:c.column for c in ws[1]}
        for cell in ws[1]:cell.font=Font(bold=True,color='FFFFFF');cell.fill=PatternFill('solid',fgColor='1F4E78')
        for header in date_headers.intersection(headers):
            for row in range(2,ws.max_row+1):ws.cell(row,headers[header]).number_format='yyyy-mm-dd'
        for col in ws.columns:ws.column_dimensions[col[0].column_letter].width=min(55,max(12,max(len(str(c.value)) if c.value is not None else 0 for c in col)+2))
    wb.save(path)
def self_test():
    dates=pd.bdate_range('2020-01-01',periods=300);close=100+np.sin(np.arange(300)/8)*10;df=pd.DataFrame({'Open':close-.2,'High':close+1,'Low':close-1,'Close':close,'Volume':1000000},index=dates);assert 'ATR' in indicators(df);print('SCANNER SELF-TEST PASSED')
def main():
    parser=argparse.ArgumentParser();parser.add_argument('--self-test',action='store_true');parser.add_argument('--out',default='reports/Hybrid_Paper_Scanner_V3.xlsx');parser.add_argument('--refresh-sp500',action='store_true');args=parser.parse_args()
    if args.self_test:self_test();return
    cfg=json.loads((ROOT/'strategy_config.json').read_text(encoding='utf-8'));universe,audit,issues=build_universe(ROOT,cfg['universe'],refresh=args.refresh_sp500);data=download(universe.Symbol.tolist(),cfg['universe']['minimum_history_bars'],issues);meta=universe.set_index('Symbol').to_dict('index');rows=[];no_setup=[];raw_total=score_total=0
    for symbol,df in data.items():
        signal,raw,passed=scan_one(df,cfg);raw_total+=raw;score_total+=passed;m=meta.get(symbol,{})
        if signal is None:
            ind=indicators(df);no_setup.append({'Symbol':symbol,'Company':m.get('Company',symbol),'Universe':m.get('Universe',''),'Latest Close':round(float(ind.Close.iloc[-1]),4),'RSI14':round(float(ind.RSI.iloc[-1]),2),'EMA9':round(float(ind.EMA9.iloc[-1]),4),'EMA20':round(float(ind.EMA20.iloc[-1]),4),'No Setup Reason':'No active divergence scoring >= 60 within 20 bars','Data Status':'OK'});continue
        event_id=hashlib.sha256(f"{symbol}|{signal['direction']}|{pd.Timestamp(signal['pivot2_date']).date()}|{cfg['strategy_id']}".encode()).hexdigest()[:16]
        row={'Event ID':event_id,'Symbol':symbol,'Company':m.get('Company',symbol),'Universe':m.get('Universe',''),'Sector':m.get('Sector',''),'Direction':signal['direction'],'Score':signal['score'],'Status':signal['status'],'Signal Date':pd.Timestamp(signal['signal_date']),'Signal Age':signal['age'],'Search Bars Remaining':20-signal['age'],'Trigger Delay':signal['trigger_delay'],'Quality':signal['quality'],'Latest Close':round(signal['latest_close'],4),'RSI14':round(signal['rsi'],2),'EMA9':round(signal['ema9'],4),'EMA20':round(signal['ema20'],4),'Main Blocker':signal['status'].replace('_',' ').title(),'Chart Link':f'https://www.tradingview.com/chart/?symbol={symbol}'}
        if signal['status']=='ENTRY_READY':row.update(trade_plan(signal,cfg))
        rows.append(row)
    rows=sorted(rows,key=lambda r:(r['Status']!='ENTRY_READY',-r['Score']));funnel={'S&P 500 Static CSV':int(audit.loc[audit.Source=='S&P 500 static CSV','Final'].iloc[0]),'Nasdaq >= $1B':int(audit.loc[audit.Source=='Nasdaq >= $1B','Final'].iloc[0]),'Final Universe':len(universe),'Valid Price Data':len(data),'Raw Divergences':raw_total,'Score Passed':score_total,'Active Events':len(rows),'Entry Ready':sum(r['Status']=='ENTRY_READY' for r in rows),'Waiting':sum(r['Status'].startswith('WAIT') for r in rows),'No Setup':len(no_setup),'Data Issues':len(issues)}
    output=ROOT/args.out;write_report(output,rows,no_setup,audit,universe,issues,funnel);print(f'Created {output}; S&P500={funnel["S&P 500 Static CSV"]}; final={len(universe)}; ready={funnel["Entry Ready"]}')
if __name__=='__main__':main()
